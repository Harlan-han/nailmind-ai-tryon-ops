"""Sync curated sample image assets into local uploads.

This script only downloads files. It does not read or mutate the database.
"""
from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path

import httpx
from openpyxl import load_workbook


MIN_IMAGE_BYTES = 1024


@dataclass(frozen=True)
class AssetItem:
    category: str
    source_url: str
    target_path: Path


@dataclass
class SyncStats:
    total: int = 0
    downloaded: int = 0
    would_download: int = 0
    skipped: int = 0
    failed: int = 0


def find_default_excel_file(project_root: Path) -> Path:
    excel_files = sorted(
        path
        for path in project_root.glob("*.xlsx")
        if not path.name.startswith("~$")
    )
    if not excel_files:
        raise FileNotFoundError(f"No .xlsx file found in {project_root}")
    return excel_files[0]


def build_asset_plan(excel_path: Path, uploads_dir: Path) -> list[AssetItem]:
    workbook = load_workbook(excel_path, data_only=True)
    if "款式图" not in workbook.sheetnames:
        raise ValueError("Workbook is missing sheet: 款式图")
    if "手图" not in workbook.sheetnames:
        raise ValueError("Workbook is missing sheet: 手图")

    assets: list[AssetItem] = []

    designs_sheet = workbook["款式图"]
    for row in designs_sheet.iter_rows(min_row=2, values_only=True):
        design_id, original_url, enhanced_url = row[:3]
        if not design_id:
            continue
        design_number = int(design_id)
        if original_url:
            assets.append(
                AssetItem(
                    category="design_original",
                    source_url=str(original_url).strip(),
                    target_path=uploads_dir / "designs" / "originals" / f"design_{design_number:02d}.jpg",
                )
            )
        if enhanced_url:
            assets.append(
                AssetItem(
                    category="design_cover",
                    source_url=str(enhanced_url).strip(),
                    target_path=uploads_dir / "designs" / f"design_{design_number:02d}.jpg",
                )
            )

    hands_sheet = workbook["手图"]
    for index, row in enumerate(hands_sheet.iter_rows(min_row=2, values_only=True), start=1):
        hand_url = row[0] if row else None
        if not hand_url:
            continue
        assets.append(
            AssetItem(
                category="hand_photo",
                source_url=str(hand_url).strip(),
                target_path=uploads_dir / "hands" / f"hand_{index:02d}.jpg",
            )
        )

    return assets


def needs_download(path: Path, force: bool, min_bytes: int = MIN_IMAGE_BYTES) -> bool:
    return force or not path.is_file() or path.stat().st_size < min_bytes


def download_asset(client: httpx.Client, asset: AssetItem, force: bool, dry_run: bool) -> str:
    if not needs_download(asset.target_path, force):
        return "skipped"

    if dry_run:
        return "would_download"

    asset.target_path.parent.mkdir(parents=True, exist_ok=True)
    response = client.get(asset.source_url)
    response.raise_for_status()
    if len(response.content) < MIN_IMAGE_BYTES:
        raise ValueError(f"Downloaded file is too small: {len(response.content)} bytes")

    temp_path = asset.target_path.with_name(f"{asset.target_path.name}.tmp")
    temp_path.write_bytes(response.content)
    temp_path.replace(asset.target_path)
    return "downloaded"


def sync_assets(assets: list[AssetItem], timeout: float, force: bool = False, dry_run: bool = False) -> dict[str, SyncStats]:
    stats: dict[str, SyncStats] = {}
    headers = {"User-Agent": "NailMind official asset sync/1.0"}
    with httpx.Client(timeout=timeout, follow_redirects=True, headers=headers) as client:
        for asset in assets:
            category_stats = stats.setdefault(asset.category, SyncStats())
            category_stats.total += 1
            try:
                result = download_asset(client, asset, force=force, dry_run=dry_run)
            except Exception as exc:  # pragma: no cover - printed for manual operations
                category_stats.failed += 1
                print(f"[failed] {asset.category} {asset.target_path.name}: {exc}")
                continue
            if result == "would_download":
                category_stats.would_download += 1
                print(f"[would-download] {asset.category} {asset.target_path}")
            elif result == "downloaded":
                category_stats.downloaded += 1
                print(f"[downloaded] {asset.category} {asset.target_path}")
            else:
                category_stats.skipped += 1

    return stats


def parse_args() -> argparse.Namespace:
    backend_dir = Path(__file__).resolve().parent
    project_root = backend_dir.parent.parent
    default_excel = find_default_excel_file(project_root)

    parser = argparse.ArgumentParser(description="Download curated NailMind sample assets into backend/uploads.")
    parser.add_argument("--excel", type=Path, default=default_excel)
    parser.add_argument("--uploads-dir", type=Path, default=backend_dir / "uploads")
    parser.add_argument("--timeout", type=float, default=30.0)
    parser.add_argument("--force", action="store_true", help="Re-download files even when local copies look valid.")
    parser.add_argument("--dry-run", action="store_true", help="Print what would be downloaded without writing files.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    assets = build_asset_plan(args.excel, args.uploads_dir)
    stats = sync_assets(assets, timeout=args.timeout, force=args.force, dry_run=args.dry_run)

    print("\nSummary")
    for category in sorted(stats):
        item = stats[category]
        print(
            f"{category}: total={item.total}, downloaded={item.downloaded}, "
            f"would_download={item.would_download}, skipped={item.skipped}, failed={item.failed}"
        )

    return 1 if any(item.failed for item in stats.values()) else 0


if __name__ == "__main__":
    raise SystemExit(main())
